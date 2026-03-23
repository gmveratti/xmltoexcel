import os
import shutil
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from typing import List, Dict, Any

import pandas as pd
import rarfile
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm

# --- Configuration ---
XML_NAMESPACE = {"ns": "http://www.portalfiscal.inf.br/cte"}

EXCEL_HEADERS: List[str] = [
    "(CTe)", "chv_cte_Id", "(ide)", "ide_cUF", "ide_cCT", "ide_CFOP", "ide_natOp",
    "ide_mod", "ide_serie", "ide_nCT", "ide_dhEmi", "ide_tpImp", "ide_tpEmis",
    "ide_cDV", "ide_tpAmb", "ide_tpCTe", "ide_procEmi", "ide_verProc",
    "ide_cMunEnv", "ide_xMunEnv", "ide_UFEnv", "ide_modal", "ide_tpServ",
    "ide_cMunIni", "ide_xMunIni", "ide_UFIni", "ide_cMunFim", "ide_xMunFim",
    "ide_UFFim", "ide_retira", "ide_indIEToma", "(ide_toma3)", "ide_toma3_toma",
    "(ide_toma4)", "ide_toma4_toma", "ide_toma4_CNPJ", "ide_toma4_IE",
    "ide_toma4_xNome", "ide_toma4_xFant", "ide_toma4_fone",
    "(ide_toma4_enderToma)", "ide_toma4_enderToma_xLgr",
    "ide_toma4_enderToma_nro", "ide_toma4_enderToma_xBairro",
    "ide_toma4_enderToma_cMun", "ide_toma4_enderToma_xMun",
    "ide_toma4_enderToma_CEP", "ide_toma4_enderToma_UF",
    "ide_toma4_enderToma_cPais", "ide_toma4_enderToma_xPais", "ide_toma4_email",
    "(compl)", "(compl_Entrega)", "compl_Entrega_tpPer", "compl_Entrega_dProg",
    "compl_Entrega_tpHor", "compl_origCalc", "compl_destCalc", "compl_xObs",
    "(emit)", "emit_CNPJ", "emit_IE", "emit_xNome", "emit_xFant",
    "(emit_enderEmit)", "emit_enderEmit_xLgr",
    "emit_enderEmit_nro", "emit_enderEmit_xCpl", "emit_enderEmit_xBairro",
    "emit_enderEmit_cMun", "emit_enderEmit_xMun", "emit_enderEmit_CEP",
    "emit_enderEmit_UF", "emit_enderEmit_fone", "(rem)", "rem_CNPJ", "rem_CPF",
    "rem_IE", "rem_xNome", "rem_xFant", "rem_fone", "(rem_enderReme)",
    "rem_enderReme_xLgr", "rem_enderReme_nro", "rem_enderReme_xBairro",
    "rem_enderReme_cMun", "rem_enderReme_xMun", "rem_enderReme_CEP",
    "rem_enderReme_UF", "rem_enderReme_cPais", "rem_enderReme_xPais",
    "rem_email", "(exped)", "exped_CNPJ", "exped_IE", "exped_xNome",
    "exped_fone", "(exped_enderExped)", "exped_enderExped_xLgr",
    "exped_enderExped_nro", "exped_enderExped_xBairro",
    "exped_enderExped_cMun", "exped_enderExped_xMun", "exped_enderExped_CEP",
    "exped_enderExped_UF", "exped_enderExped_cPais", "exped_enderExped_xPais",
    "exped_email", "(receb)", "receb_CNPJ", "receb_CPF", "receb_IE", "receb_xNome",
    "receb_fone", "(receb_enderReceb)", "receb_enderReceb_xLgr",
    "receb_enderReceb_nro", "receb_enderReceb_xBairro",
    "receb_enderReceb_cMun", "receb_enderReceb_xMun", "receb_enderReceb_CEP",
    "receb_enderReceb_UF", "receb_enderReceb_cPais", "receb_enderReceb_xPais",
    "receb_email", "(dest)", "dest_CNPJ", "dest_CPF", "dest_IE", "dest_xNome",
    "dest_fone", "(dest_enderDest)", "dest_enderDest_xLgr",
    "dest_enderDest_nro", "dest_enderDest_xBairro", "dest_enderDest_cMun",
    "dest_enderDest_xMun", "dest_enderDest_CEP", "dest_enderDest_UF",
    "dest_enderDest_cPais", "dest_enderDest_xPais", "dest_email", "(vPrest)",
    "vPrest_vTPrest", "vPrest_vRec", "(vPrest_Comp)", "vPrest_xNome", "vPrest_vComp",
    "(imp)", "imp_CST", "imp_vBC", "imp_pICMS", "imp_vICMS",
    "imp_vBCSTRet", "imp_vICMSSTRet", "imp_pICMSSTRet", "(imp_ICMSOutraUF)",
    "imp_ICMSOutraUF_CST", "imp_ICMSOutraUF_vBCOutraUF",
    "imp_ICMSOutraUF_pICMSOutraUF", "imp_ICMSOutraUF_vICMSOutraUF",
    "imp_vTotTrib", "imp_infAdFisco", "(infCTeNorm)",
    "(infCTeNorm_infCarga)", "infCTeNorm_infCarga_vCarga",
    "infCTeNorm_infCarga_proPred", "(infCTeNorm_infCteSub)",
    "infCTeNorm_infCteSub_chCte", "infCTeNorm_infCteSub_indAlteraToma",
    "(infCTeNorm_infServVinc)", "(infCTeNorm_infServVinc_infCTeMultimodal)",
    "infCTeNorm_infServVinc_infCTeMultimodal_chCTeMultimodal",
    "(infCteComp)", "infCteComp_chCTe"
]

# Columns handled exclusively by the ICMS routing block — MUST be skipped in the dynamic loop
# Using a set for O(1) lookup performance
ICMS_COLUMNS: set = {
    "imp_CST", "imp_vBC", "imp_pICMS", "imp_vICMS",
    "imp_vBCSTRet", "imp_vICMSSTRet", "imp_pICMSSTRet",
    "imp_ICMSOutraUF_CST", "imp_ICMSOutraUF_vBCOutraUF",
    "imp_ICMSOutraUF_pICMSOutraUF", "imp_ICMSOutraUF_vICMSOutraUF",
}

MANUAL_COLUMNS: set = {"chv_cte_Id", "vPrest_xNome", "vPrest_vComp"}


class ArchiveHandler:
    """Handles extraction of .rar and .zip archives."""

    def __init__(self, archive_path: str):
        if not os.path.exists(archive_path):
            raise FileNotFoundError(f"Archive not found at: {archive_path}")
        self.archive_path = archive_path
        self.temp_dir = tempfile.mkdtemp(prefix="cte_extraction_")

    def _extract_recursive(self):
        while True:
            archives_found = False
            for root, _, files in os.walk(self.temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    extract_path = os.path.dirname(file_path)
                    try:
                        if file.lower().endswith(".rar"):
                            with rarfile.RarFile(file_path) as rf:
                                rf.extractall(path=extract_path)
                            os.remove(file_path)
                            archives_found = True
                        elif file.lower().endswith(".zip"):
                            with zipfile.ZipFile(file_path) as zf:
                                zf.extractall(path=extract_path)
                            os.remove(file_path)
                            archives_found = True
                    except Exception:
                        os.rename(file_path, file_path + ".failed")
            if not archives_found:
                break

    def extract_all(self):
        print(f"Extracting main archive: {os.path.basename(self.archive_path)}...")
        try:
            with rarfile.RarFile(self.archive_path) as rf:
                rf.extractall(path=self.temp_dir)
            print("Scanning for nested archives...")
            self._extract_recursive()
        except rarfile.Error as e:
            raise RuntimeError(
                f"Failed to extract RAR file. Ensure 'unrar' is installed. Details: {e}"
            )

    def find_xml_files(self) -> List[str]:
        xml_files = []
        for root, _, files in os.walk(self.temp_dir):
            for file in files:
                if file.lower().endswith(".xml"):
                    xml_files.append(os.path.join(root, file))
        return xml_files

    def cleanup(self):
        if os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
            except OSError as e:
                print(f"Warning: Could not clean up temp directory {self.temp_dir}: {e}")


class XMLParser:
    """Parses CT-e XML com mapeamento direto e explícito (Sem helpers complexos)."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.ns = {"ns": "http://www.portalfiscal.inf.br/cte"}
        self.tree = ET.parse(file_path)
        self.root = self.tree.getroot()

    def extract_data(self, headers: List[str]) -> List[Dict[str, str]]:
        base_data = {header: "" for header in headers}
        
        inf_cte_node = self.root.find(".//ns:infCte", self.ns)
        if inf_cte_node is None:
            return [base_data]

        base_data["chv_cte_Id"] = inf_cte_node.get("Id", "").replace("CTe", "")

        # --- 1. EXTRAÇÃO DINÂMICA SIMPLES (Ignora os Impostos Específicos) ---
        skip_cols = {
            "imp_CST", "imp_vBC", "imp_pICMS", "imp_vICMS",
            "imp_vBCSTRet", "imp_vICMSSTRet", "imp_pICMSSTRet",
            "imp_ICMSOutraUF_CST", "imp_ICMSOutraUF_vBCOutraUF",
            "imp_ICMSOutraUF_pICMSOutraUF", "imp_ICMSOutraUF_vICMSOutraUF",
            "chv_cte_Id", "vPrest_xNome", "vPrest_vComp"
        }
        
        for header in headers:
            if header.startswith("(") or header in skip_cols:
                continue
            xpath = ".//" + "/".join([f"ns:{p}" for p in header.split("_")])
            element = inf_cte_node.find(xpath, self.ns)
            if element is not None and element.text:
                base_data[header] = element.text.strip()

        # --- 2. EXTRAÇÃO DIRETA DE ICMS (IMPOSSÍVEL DE FALHAR) ---
        icms_node = inf_cte_node.find(".//ns:ICMS", self.ns)
        if icms_node is not None and len(icms_node) > 0:
            grupo_icms = icms_node[0] # Pega a tag do grupo (ex: ICMS60, ICMSOutraUF)
            tag_grupo = grupo_icms.tag.split("}")[-1].upper()

            if tag_grupo == "ICMSOUTRAUF":
                for child in grupo_icms:
                    c_tag = child.tag.split("}")[-1]
                    if c_tag == "CST": base_data["imp_ICMSOutraUF_CST"] = child.text.strip() if child.text else ""
                    elif c_tag == "vBCOutraUF": base_data["imp_ICMSOutraUF_vBCOutraUF"] = child.text.strip() if child.text else ""
                    elif c_tag == "pICMSOutraUF": base_data["imp_ICMSOutraUF_pICMSOutraUF"] = child.text.strip() if child.text else ""
                    elif c_tag == "vICMSOutraUF": base_data["imp_ICMSOutraUF_vICMSOutraUF"] = child.text.strip() if child.text else ""

            elif tag_grupo == "ICMS60":
                for child in grupo_icms:
                    c_tag = child.tag.split("}")[-1]
                    if c_tag == "CST": base_data["imp_CST"] = child.text.strip() if child.text else ""
                    elif c_tag == "vBCSTRet": base_data["imp_vBCSTRet"] = child.text.strip() if child.text else ""
                    elif c_tag == "vICMSSTRet": base_data["imp_vICMSSTRet"] = child.text.strip() if child.text else ""
                    elif c_tag == "pICMSSTRet": base_data["imp_pICMSSTRet"] = child.text.strip() if child.text else ""

            else:
                # ICMS00, ICMS20, ICMS90, etc.
                for child in grupo_icms:
                    c_tag = child.tag.split("}")[-1]
                    if c_tag == "CST": base_data["imp_CST"] = child.text.strip() if child.text else ""
                    elif c_tag == "vBC": base_data["imp_vBC"] = child.text.strip() if child.text else ""
                    elif c_tag == "pICMS": base_data["imp_pICMS"] = child.text.strip() if child.text else ""
                    elif c_tag == "vICMS": base_data["imp_vICMS"] = child.text.strip() if child.text else ""
                    elif c_tag == "vBCSTRet": base_data["imp_vBCSTRet"] = child.text.strip() if child.text else ""
                    elif c_tag == "vICMSSTRet": base_data["imp_vICMSSTRet"] = child.text.strip() if child.text else ""
                    elif c_tag == "pICMSSTRet": base_data["imp_pICMSSTRet"] = child.text.strip() if child.text else ""

        # --- 3. DUPLICAÇÃO DE COMPONENTES ---
        rows = []
        comps = inf_cte_node.findall(".//ns:vPrest/ns:Comp", self.ns)
        
        if comps:
            for comp in comps:
                row = base_data.copy()
                x_nome = comp.find("ns:xNome", self.ns)
                v_comp = comp.find("ns:vComp", self.ns)
                if x_nome is not None and x_nome.text: row["vPrest_xNome"] = x_nome.text.strip()
                if v_comp is not None and v_comp.text: row["vPrest_vComp"] = v_comp.text.strip()
                rows.append(row)
        else:
            rows.append(base_data.copy())

        return rows


class ExcelExporter:
    """Exports parsed data to a styled Excel file."""

    def __init__(self, data: List[Dict[str, Any]], headers: List[str]):
        self.data = data
        self.headers = headers

    def export(self, output_filename: str):
        if not self.data:
            return
        print(f"Generating Excel file: {output_filename}...")

        df = pd.DataFrame(self.data)
        df = df.reindex(columns=self.headers, fill_value="")

        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="CTe Data")
            ws = writer.sheets["CTe Data"]

            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            bold_font = Font(bold=True)

            for col_idx, header_text in enumerate(self.headers, 1):
                ws.cell(row=1, column=col_idx).font = bold_font
                if header_text.startswith("(") and header_text.endswith(")"):
                    for row in ws.iter_rows(
                        min_row=1, max_row=ws.max_row,
                        min_col=col_idx, max_col=col_idx
                    ):
                        for cell in row:
                            cell.fill = gray_fill


class AppController:
    """Main orchestrator."""

    def __init__(self, input_dir: str = "files"):
        self.input_dir = input_dir
        if not os.path.exists(self.input_dir):
            os.makedirs(self.input_dir)
            print(f"Created directory: '{self.input_dir}/'. Please place your .rar inside it.")
            exit()

    def run(self):
        print("--- CT-e XML to Excel Converter ---")
        rar_filename = input(f"Enter the name of the .rar file (must be in '{self.input_dir}/'): ").strip()
        rar_path = os.path.join(self.input_dir, rar_filename)

        if not os.path.exists(rar_path):
            print(f"\nError: File not found at '{rar_path}'.")
            return

        archive_handler = ArchiveHandler(rar_path)

        try:
            archive_handler.extract_all()
            xml_files = archive_handler.find_xml_files()
            print(f"Total XML files found: {len(xml_files)}")

            all_data: List[Dict[str, Any]] = []
            error_details: List[str] = []

            for xml_file in tqdm(xml_files, desc="Processing XMLs", unit="file"):
                try:
                    parser = XMLParser(xml_file)
                    rows = parser.extract_data(EXCEL_HEADERS)
                    all_data.extend(rows)
                except Exception as e:
                    file_name = os.path.basename(xml_file)
                    error_details.append(f"Arquivo: [{file_name}] | Erro: {str(e)}")

            if error_details:
                log_filename = f"erros_{os.path.splitext(rar_filename)[0]}.log"
                with open(log_filename, "w", encoding="utf-8") as f:
                    f.write("RELATÓRIO DE ERROS DE PROCESSAMENTO - CT-e\n")
                    f.write("=" * 50 + "\n")
                    for err in error_details:
                        f.write(err + "\n")
                print(f"\nAviso: {len(error_details)} arquivo(s) XML falharam.")
                print(f"Verifique '{log_filename}' para detalhes.")

            output_filename = f"{os.path.splitext(rar_filename)[0]}.xlsx"
            exporter = ExcelExporter(all_data, EXCEL_HEADERS)
            exporter.export(output_filename)
            print(f"\nSucesso! Excel salvo como: {output_filename}")

        except Exception as e:
            print(f"\nErro inesperado no processo principal: {e}")
        finally:
            print("\nLimpando arquivos temporários...")
            archive_handler.cleanup()


if __name__ == "__main__":
    controller = AppController()
    controller.run()