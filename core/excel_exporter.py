# core/excel_exporter.py

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font, Alignment

from core.constants import (
    GRAY_FILL_COLOR, ACCOUNTING_FORMAT,
    EVENT_SHEET_HEADERS, EVENT_DETAIL_COL_WIDTH, EVENT_KEY_COL_WIDTH
)
from core.strategy import resolve_strategy, DocumentStrategy

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta para Excel em fluxo de disco (O(1) memória RAM)."""

    def __init__(self, main_data: List[Dict[str, Any]], 
                 strategy: DocumentStrategy,
                 event_data: Optional[List[Dict[str, Any]]] = None):
        self.main_data = main_data
        self.event_data = event_data or []
        self.strategy = strategy

        # Fills reutilizáveis
        self._gray_fill = PatternFill(
            start_color=GRAY_FILL_COLOR, end_color=GRAY_FILL_COLOR, fill_type="solid"
        )
        self._bold_font = Font(bold=True)
        self._center_align = Alignment(horizontal="center", vertical="center")

    def export(self, output_filename: str):
        if not self.main_data and not self.event_data:
            raise ValueError(f"Nenhum dado {self.strategy.doc_type_name} válido para exportar.")

        wb = openpyxl.Workbook(write_only=True)

        sheet_title = "NF-e Data" if self.strategy.doc_type_name == "NFE" else "CTe Data"
        ws_main = wb.create_sheet(title=sheet_title)
        self._build_main_sheet(ws_main)

        if self.event_data:
            ws_events = wb.create_sheet(title="Eventos e Correções")
            self._build_events_sheet(ws_events)

        wb.save(output_filename)
        logger.info("Excel salvo em: %s", output_filename)

    def _build_main_sheet(self, ws):
        """Constrói a aba principal (CTE ou NFE) de forma genérica via estratégia."""
        headers = self.strategy.excel_headers
        accounting_columns = self.strategy.accounting_cols

        # Colunas dinâmicas (ex: comp_* no CTe)
        dynamic_cols = sorted({
            key for row in self.main_data
            for key in row
            if key.startswith("comp_") and key not in headers
        })

        final_headers = headers + dynamic_cols

        for idx, _ in enumerate(final_headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 25

        # Cabeçalho
        header_cells = []
        for header in final_headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = self._bold_font
            cell.alignment = self._center_align
            if self.strategy.is_gray_col(header):
                cell.fill = self._gray_fill
            header_cells.append(cell)
        ws.append(header_cells)

        # Linhas de dados
        for row_data in self.main_data:
            row_cells = []
            for header in final_headers:
                raw_val = row_data.get(header, "")
                raw_val = str(raw_val).strip() if raw_val is not None else ""

                cell = WriteOnlyCell(ws)
                cell.alignment = self._center_align

                if self.strategy.is_gray_col(header):
                    cell.fill = self._gray_fill
                    row_cells.append(cell)
                    continue

                # Lógica de tipos (Data, Monetário, Texto)
                if (header == "ide_dhEmi" or header == "dhEmi") and raw_val:
                    try:
                        dt_obj = datetime.fromisoformat(raw_val).replace(tzinfo=None)
                        cell.value = dt_obj
                        cell.number_format = "DD/MM/YYYY HH:mm:SS"
                    except ValueError:
                        cell.value = raw_val
                        cell.number_format = "@"
                elif header in accounting_columns or header.startswith("comp_"):
                    if raw_val:
                        try:
                            cell.value = float(raw_val)
                            cell.number_format = ACCOUNTING_FORMAT
                        except ValueError:
                            cell.value = raw_val
                            cell.number_format = "@"
                else:
                    cell.value = raw_val
                    cell.number_format = "@"

                row_cells.append(cell)
            ws.append(row_cells)

    def _build_events_sheet(self, ws):
        """Aba de eventos (comum a todos os tipos)."""
        ws.column_dimensions["A"].width = EVENT_KEY_COL_WIDTH
        ws.column_dimensions["D"].width = EVENT_DETAIL_COL_WIDTH

        header_cells = []
        for header in EVENT_SHEET_HEADERS:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = self._bold_font
            cell.alignment = self._center_align
            header_cells.append(cell)
        ws.append(header_cells)

        for row_data in self.event_data:
            row_cells = []
            for header in EVENT_SHEET_HEADERS:
                raw_val = row_data.get(header, "")
                raw_val = str(raw_val).strip() if raw_val is not None else ""

                cell = WriteOnlyCell(ws, value=raw_val)
                cell.alignment = self._center_align
                cell.number_format = "@"
                row_cells.append(cell)
            ws.append(row_cells)
