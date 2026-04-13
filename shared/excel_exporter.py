# core/excel_exporter.py

import logging
from typing import List, Dict, Any
from datetime import datetime

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font, Alignment

from shared.constants import (
    GRAY_FILL_COLOR, ACCOUNTING_FORMAT, MAX_COLUMN_WIDTH,
    EVENT_SHEET_HEADERS, EVENT_DETAIL_COL_WIDTH, EVENT_KEY_COL_WIDTH,
    NFE_ACCOUNTING_COLUMNS
)
from shared.models import DocType

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Exporta para Excel em fluxo de disco (O(1) memória RAM). Suporta CT-e e NF-e."""

    def __init__(self, main_data: List[Dict[str, Any]], main_headers: List[str],
                 event_data: List[Dict[str, Any]] = None, doc_type: DocType = DocType.CTE):
        self.main_data = main_data
        self.main_headers = main_headers
        self.event_data = event_data or []
        self.doc_type = doc_type

    def export(self, output_filename: str):
        if not self.main_data and not self.event_data:
            raise ValueError("Nenhum dado válido para exportar.")

        wb = openpyxl.Workbook(write_only=True)

        sheet_title = "CTe Data" if self.doc_type == DocType.CTE else "NFe Data"
        ws_main = wb.create_sheet(title=sheet_title)
        self._build_main_sheet(ws_main)

        if self.event_data:
            ws_events = wb.create_sheet(title="Eventos e Correções")
            self._build_events_sheet(ws_events)

        wb.save(output_filename)
        logger.info("Excel salvo em: %s", output_filename)

    def _build_main_sheet(self, ws):
        gray_fill = PatternFill(start_color=GRAY_FILL_COLOR, end_color=GRAY_FILL_COLOR, fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Contabilidade do CT-e
        cte_accounting = {
            "vPrest_vTPrest", "vPrest_vRec", "imp_vBC", "imp_pICMS",
            "imp_vICMS", "imp_vBCSTRet", "imp_vICMSSTRet", "imp_pICMSSTRet",
            "imp_ICMSOutraUF_vBCOutraUF", "imp_ICMSOutraUF_pICMSOutraUF",
            "imp_ICMSOutraUF_vICMSOutraUF", "imp_vTotTrib", 
            "comp_FRETE_VALOR", "comp_IMPOSTOS", "comp_PEDAGIO",
            "comp_VALOR_FRETE", "comp_VALOR_ICMS", "comp_VALOR_PEDAGIO"
        }

        # Fundir cabeçalhos dinâmicos apenas se for CT-e
        dynamic_cols = []
        if self.doc_type == DocType.CTE:
            dynamic_cols = sorted({
                key for row in self.main_data
                for key in row
                if key.startswith("comp_") and key not in self.main_headers
            })

        final_headers = self.main_headers + dynamic_cols

        # Larguras padrão
        for idx, _ in enumerate(final_headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = 25

        # --- Cabeçalho ---
        header_cells = []
        for header in final_headers:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            if header.startswith("(") and header.endswith(")"):
                cell.fill = gray_fill
            header_cells.append(cell)
        ws.append(header_cells)

        # --- Linhas de Dados ---
        for row_data in self.main_data:
            row_cells = []
            for header in final_headers:
                raw_val = row_data.get(header, "")
                raw_val = str(raw_val).strip() if raw_val is not None else ""
                
                cell = WriteOnlyCell(ws)
                cell.alignment = center_alignment

                if header.startswith("(") and header.endswith(")"):
                    cell.fill = gray_fill
                    row_cells.append(cell)
                    continue

                if header in ("ide_dhEmi", "Data do Evento") and raw_val:
                    try:
                        dt_obj = datetime.fromisoformat(raw_val).replace(tzinfo=None)
                        cell.value = dt_obj
                        cell.number_format = 'DD/MM/YYYY HH:MM:SS'
                    except ValueError:
                        cell.value = raw_val
                        cell.number_format = '@'
                
                # Validação Híbrida de Moeda (Suporta ambos os documentos perfeitamente)
                elif (self.doc_type == DocType.CTE and (header in cte_accounting or header.startswith("comp_"))) or \
                     (self.doc_type == DocType.NFE and header in NFE_ACCOUNTING_COLUMNS):
                    if raw_val:
                        try:
                            cell.value = float(raw_val)
                            cell.number_format = ACCOUNTING_FORMAT
                        except ValueError:
                            cell.value = raw_val
                            cell.number_format = '@'
                else:
                    cell.value = raw_val
                    cell.number_format = '@'
                
                row_cells.append(cell)
            ws.append(row_cells)

    def _build_events_sheet(self, ws):
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = EVENT_KEY_COL_WIDTH
        ws.column_dimensions['D'].width = EVENT_DETAIL_COL_WIDTH

        header_cells = []
        for header in EVENT_SHEET_HEADERS:
            cell = WriteOnlyCell(ws, value=header)
            cell.font = bold_font
            cell.alignment = center_alignment
            header_cells.append(cell)
        ws.append(header_cells)

        for row_data in self.event_data:
            row_cells = []
            for header in EVENT_SHEET_HEADERS:
                raw_val = row_data.get(header, "")
                raw_val = str(raw_val).strip() if raw_val is not None else ""
                
                cell = WriteOnlyCell(ws, value=raw_val)
                cell.alignment = center_alignment
                cell.number_format = '@'
                row_cells.append(cell)
            ws.append(row_cells)
